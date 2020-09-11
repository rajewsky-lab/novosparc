#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Thu Sep 10 12:09:06 2020

@author: mornitzan
"""



#%%


#%%
from openpyxl import load_workbook
import pandas as pd
from novosparc import *
import scipy as sc

#%%


wb = load_workbook(filename='corti_expression.xlsx', read_only=True)
ws = wb['mastertable.csv']

# Read the cell values into a list of lists
data_rows = []
for row in ws['I4':'GR880']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

# Transform into dataframe
df = pd.DataFrame(data_rows)
expression_matrix = pd.DataFrame.as_matrix(df)

data_rows = []
for row in ws['I3':'GR3']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

# Transform into dataframe
df = pd.DataFrame(data_rows)
gene_names = df.astype(str).values.tolist()

data_rows = []
for row in ws['H4':'H880']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

# Transform into dataframe
df = pd.DataFrame(data_rows)
subpopulations = df.astype(str).values.tolist()
subpopulations = [item for sublist in subpopulations for item in sublist]

u, indices = np.unique(subpopulations, return_inverse=True)

data_rows = []
for row in ws['F4':'F880']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

# Transform into dataframe
df = pd.DataFrame(data_rows)
anatomical_origin = df.astype(str).values.tolist()
anatomical_origin = [item for sublist in anatomical_origin for item in sublist]
u_anat, indices_anat = np.unique(anatomical_origin, return_inverse=True)
     
#%%
dge_full = np.copy(expression_matrix) #cells by genes

num_cells = dge_full.shape[0]
num_genes = dge_full.shape[1]

ind_type = np.copy(indices) #10 cell types
names_type = np.copy(u) #10 cell types

gene_names_flat = [item for sublist in gene_names for item in sublist]

#ind_loc = np.copy(indices_anat) 
ind_loc = np.zeros(num_cells)
ind_loc[np.argwhere(indices_anat==1)] = 0 #base
ind_loc[np.argwhere(indices_anat==2)] = 1 #middle
ind_loc[np.argwhere(indices_anat==0)] = 2 #apex
ind_loc[np.argwhere(indices_anat==3)] = np.nan #nan
names_loc = [u_anat[1],u_anat[2],u_anat[0],u_anat[3]]

cell_filter = np.array(np.argwhere(~np.isnan(ind_loc)).flatten())

dge_full = dge_full[cell_filter,:]
num_cells = dge_full.shape[0]
ind_type = ind_type[cell_filter]
ind_loc = ind_loc[cell_filter]


#%%
########################################################################

#%%

gene_names_flat = [item for sublist in gene_names for item in sublist]
gene_names_flat = np.array(gene_names_flat)

#%%

cur_filter = np.argwhere(ind_type==7).flatten() #OHC 

locations_original = ind_loc[cur_filter]
len_locations_original = len(np.unique(ind_loc))
grid_len = 10


################################
# 2. Set the target space grid #
################################

locations = np.vstack((range(grid_len),np.ones(grid_len))).T

######################################
# 3. Setup for the OT reconstruction #
######################################

gene_count1_norm = np.copy(dge_full[cur_filter,:].T)
gene_count1_norm = gene_count1_norm / np.nansum(gene_count1_norm,axis=0)
gene_count1_norm = np.log2((gene_count1_norm)+1)

dge = gene_count1_norm.T
num_cells = dge.shape[0]

sp_expression = sc.spatial.distance.squareform(sc.spatial.distance.pdist( dge ,'correlation'))
cost_expression = sp_expression / sp_expression.max()
cost_expression -= np.mean(cost_expression)

num_locations = len(locations)
sp_locations = np.zeros((num_locations,num_locations))
for i in range(num_locations):
    for j in range(num_locations):
        sp_locations[i,j] = np.abs(i-j)
    
cost_locations = sp_locations / sp_locations.max()
cost_locations -= np.mean(cost_locations)

# no marker genes are used
cost_marker_genes = np.ones((num_cells,len(locations)))

#############################
# 4. Spatial reconstruction #
#############################

# Distributions at target and source spaces
p_locations = ot.unif(len(locations))
p_expression = ot.unif(num_cells)

alpha_linear = 0
gw = gwa.gromov_wasserstein_adjusted_norm(cost_marker_genes, cost_expression, cost_locations,
                                          alpha_linear, p_expression, p_locations,
                                          'square_loss', epsilon=5e-4, verbose=True)

# Compute sdge
sdge = np.dot(dge_full[cur_filter,:].T, gw)

# Compute mean expression distribution over embedded zones 
mean_exp_new_dist = np.zeros((grid_len,grid_len))
for i in range(grid_len):
    indices =  np.argwhere(locations_original==i).flatten()
    temp = np.sum(gw[indices,:],axis=0)
    mean_exp_new_dist[i,:] = temp/np.sum(temp)

#%
    
fonts = 20
fonts_ticks=15

grid_len_original = len_locations_original
membership_ind = np.zeros((grid_len_original,grid_len))

for i in range(grid_len_original):
    temp = np.sum(gw[locations_original==i,:],axis=0)
    membership_ind[i,:] = temp / np.sum(temp)
 
    

plt.figure(figsize=(5,4))

membership_ind_norm = np.copy(membership_ind)
membership_ind_norm = (membership_ind_norm.T/np.max(membership_ind_norm,axis=1)).T
membership_ind_norm = membership_ind_norm[:,::-1]
plt.plot(range(grid_len),membership_ind_norm[[0,2],:].T)
plt.ylabel('Cellular membership', fontsize=fonts)
plt.xlabel('Embedded layer', fontsize=fonts)
plt.legend(('base','apex'), fontsize=fonts)
plt.yticks(fontsize=fonts_ticks)
plt.xticks(fontsize=fonts_ticks)
plt.show()    

#%%    

gene_list_base = ['Otof', 'Ptprq', 'Clrn1','Calb1', 'Calb2', 'Ocm']
gene_list_apex = ['Actb', 'Myo15', 'Pcdh15','Isl1','Dach1','Eya2','Pax2']

zonated_lst_base=[]
for gene in gene_list_base:
    zonated_lst_base = np.append(zonated_lst_base, np.argwhere(gene_names_flat == gene))
zonated_lst_base = zonated_lst_base.astype(int)

zonated_lst_apex=[]
for gene in gene_list_apex:
    zonated_lst_apex = np.append(zonated_lst_apex, np.argwhere(gene_names_flat == gene))
zonated_lst_apex = zonated_lst_apex.astype(int)



#%%
fonts_leg = 13

sdge_norm = np.copy(sdge)
sdge_norm = (sdge_norm.T/np.max(sdge_norm,axis=1)).T
sdge_norm = sdge_norm[:,::-1]

plt.figure(figsize=(10,4))

plt.subplot(1,2,1)
plt.plot(range(grid_len),sdge_norm[zonated_lst_base,:].T)
plt.xlabel('Embedded layer', fontsize=fonts)
plt.ylabel('Expression', fontsize=fonts)
plt.title('Genes zonated towards base', fontsize=fonts)
plt.legend((gene_list_base), fontsize=fonts_leg)
plt.ylim(0,1)
plt.xticks(fontsize=fonts_ticks)
plt.yticks(fontsize=fonts_ticks)

plt.subplot(1,2,2)
plt.plot(range(grid_len),sdge_norm[zonated_lst_apex,:].T)
plt.xlabel('Embedded layer', fontsize=fonts)
plt.ylabel('Expression', fontsize=fonts)
plt.legend((gene_list_apex), fontsize=fonts_leg)
plt.title('Genes zonated towards apex', fontsize=fonts)
plt.ylim(0,1)
plt.yticks(fontsize=fonts_ticks)
plt.xticks(fontsize=fonts_ticks)

plt.tight_layout()

plt.show()   



#%% additional fig : expression patterns for zonated genes in reconstructed patterns
# identifying highly zonated genes

fonts_ticks1 = 13
num_genes = sdge.shape[0]

max_layer_genes = np.zeros(num_genes)

for i in range(num_genes):
    max_layer_genes[i] = np.argmax(sdge[i,:])

pval_recon = np.zeros(num_genes)
range_here = range(sdge.shape[1])
for i in range(num_genes):
    tau, p_value1 = sc.stats.kendalltau(range_here, sdge[i,:])
    tau, p_value2 = sc.stats.kendalltau(range_here[::-1], sdge[i,:])
    pval_recon[i] = min(p_value1,p_value2)

#%%
    

locations_original_0 = np.argwhere((max_layer_genes == 0) & (np.max(sdge,axis=1) > 10**-5))
locations_original_6 = np.argwhere((max_layer_genes == np.max(max_layer_genes)) & (np.max(sdge,axis=1) > 10**-5))
locations_original_0_pval = pval_recon[locations_original_0]
locations_original_6_pval = pval_recon[locations_original_6]
pval_sorted_0 = locations_original_0[np.argsort(locations_original_0_pval.flatten())]
pval_sorted_6 = locations_original_6[np.argsort(locations_original_6_pval.flatten())]

num_genes_each = 10
zonated_lst = np.vstack((pval_sorted_0[0:num_genes_each],pval_sorted_6[0:num_genes_each])).flatten()

gene_names_array = np.array(gene_names).flatten()
zonated_lst_names=[]
for gene in zonated_lst:
    temp = gene_names_array[gene].split(';')[0]
    zonated_lst_names = np.append(zonated_lst_names, temp)

ax1 = plt.subplot2grid((1, 2), (0, 0), colspan = 15)
ax2 = plt.subplot2grid((1, 2), (0, 1), rowspan = 12)


plt.subplot(1,2,1)    
x = range(sdge.shape[1])
y = sdge[zonated_lst,:]
y = (y.T /np.max(y,axis=1).T).T
ax = plt.gca()
im = ax.imshow(y)
my_xticks = ['1','2','3','4','5','6','7','8','9','10']
plt.xticks(x, my_xticks)
plt.xlabel('Embedded layer', fontsize=fonts)
my_yticks = zonated_lst_names
plt.yticks(range(len(my_yticks)), my_yticks)
plt.yticks(fontsize=fonts_ticks1)
plt.xticks(fontsize=fonts_ticks1)
plt.yticks(fontsize=fonts_ticks1)
plt.xticks(fontsize=fonts_ticks1)


locations_original = ind_loc
grid_len = len(np.unique(locations_original))

# Compute mean dge over original zones 
dge_full_mean = np.zeros((grid_len,dge_full.shape[1]))
for i in range(grid_len):
    indices =  np.argwhere(locations_original==i).flatten()
    temp = np.mean(dge_full[indices,:],axis=0)
    dge_full_mean[i,:] = temp
dge_full_mean = dge_full_mean.T 



plt.subplot(1,2,2)    
x = range(dge_full_mean.shape[1])
y = dge_full_mean[zonated_lst,:]
y = (y.T /np.max(y,axis=1).T).T
y = y[:,::-1]
ax = plt.gca()
im = ax.imshow(y)
my_xticks = ['1','2','3']
plt.xticks(x, my_xticks)
plt.xlabel('Original layer', fontsize=fonts)
my_yticks = zonated_lst_names
plt.yticks([])
plt.xticks(fontsize=fonts_ticks1)
divider = make_axes_locatable(ax)
cax = divider.append_axes("right", size="15%", pad=0.05)
plt.colorbar(im, cax=cax)
plt.yticks(fontsize=fonts_ticks1)
plt.xticks(fontsize=fonts_ticks1)

plt.show()



#%%

sort_0 = np.argsort(dge_full_mean[:,0]/(dge_full_mean[:,2]+10**-6))
sort_2 = np.argsort(dge_full_mean[:,2]/(dge_full_mean[:,0]+10**-6))

num_genes_each = 30
zonated_lst = np.vstack((sort_0[-num_genes_each:],sort_2[-num_genes_each:])).flatten()

num_cell_types = len(np.unique(ind_type))

grid_len_original = len(np.unique(locations_original))
grid_len = 5
membership_ind = np.zeros((num_cell_types,grid_len_original,grid_len))

for k in range(num_cell_types):
    print(k)
    cur_filter = np.argwhere(ind_type==k).flatten()
    locations_original_here = ind_loc[cur_filter]
    locations = np.vstack((range(grid_len),np.ones(grid_len))).T
    
    gene_count1_norm = np.copy(dge_full[cur_filter,:].T)
    gene_count1_norm = gene_count1_norm / np.nansum(gene_count1_norm,axis=0)
    gene_count1_norm = np.log2((gene_count1_norm)+1)
    dge = gene_count1_norm[zonated_lst,:].T
    num_cells = gene_count1_norm.shape[1]
    
    sp_expression = sc.spatial.distance.squareform(sc.spatial.distance.pdist( dge ,'correlation'))
    cost_expression = sp_expression / sp_expression.max()
    cost_expression -= np.mean(cost_expression)
    
    num_locations = len(locations)
    sp_locations = np.zeros((num_locations,num_locations))
    for i in range(num_locations):
        for j in range(num_locations):
            sp_locations[i,j] = np.abs(i-j)
        
    cost_locations = sp_locations / sp_locations.max()
    cost_locations -= np.mean(cost_locations)
    
    cost_marker_genes = np.ones((num_cells,len(locations)))
    p_locations = ot.unif(len(locations))
    p_expression = ot.unif(num_cells)
    
    alpha_linear = 0
    gw = gwa.gromov_wasserstein_adjusted_norm(cost_marker_genes, cost_expression, cost_locations,
                                              alpha_linear, p_expression, p_locations,
                                              'square_loss', epsilon=5e-4, verbose=True)

    for i in range(grid_len_original):
        temp = np.sum(gw[locations_original_here==i,:],axis=0)
        membership_ind[k,i,:] = temp / np.sum(temp)
    
#%%    
        
fonts = 20
fonts_ticks=15
    
fig, ax = plt.subplots(2,5, sharex=True, sharey=True, figsize=(10,4))

for i in range(1,11):
    plt.subplot(2,5,i)
    temp1 = membership_ind[i-1,0,:]; temp1 = temp1-temp1.min(); temp1 = temp1/temp1.max()
    temp2 = membership_ind[i-1,2,:]; temp2 = temp2-temp2.min(); temp2 = temp2/temp2.max()
    temp_both = np.vstack((temp1,temp2))
    if temp1[0]<temp2[0]:
        temp_both = temp_both[:,::-1]
    if i<11:        
        plt.plot(range(grid_len),temp_both.T)
    
    plt.title(names_type[i-1],fontsize=fonts)
    plt.yticks(fontsize=fonts_ticks)
    plt.xticks(fontsize=fonts_ticks)
    if i==1 or i==6:
        pass
    else:
        plt.yticks([])
    if i<6 :
        plt.xticks([])
fig.text(0.5, -0.1, 'Embedded layer', ha='center',fontsize=fonts)
fig.text(-0.05, 0.5, 'Cellular membership', va='center', rotation='vertical',fontsize=fonts)
plt.tight_layout(pad=0.0,w_pad=0.0, h_pad=1.0)
plt.show()

    


